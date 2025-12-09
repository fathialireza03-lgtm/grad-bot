# bot.py
import os
from telegram import Update, ReplyKeyboardMarkup
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    filters,
    ContextTypes,
    ConversationHandler,
)
from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv

load_dotenv()  # بارگذاری .env

# مراحل مکالمه
NAME, STUDENT_ID, GUEST_COUNT, EDIT_CONFIRM, EDIT_NAME, EDIT_GUEST = range(6)

FILE_NAME = "graduation_data.xlsx"

def init_excel():
    if not os.path.exists(FILE_NAME):
        wb = Workbook()
        ws = wb.active
        ws.append(["نام", "کد دانشجویی", "تعداد همراهان"])
        wb.save(FILE_NAME)

def find_student(student_id):
    wb = load_workbook(FILE_NAME)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if str(row[1].value) == str(student_id):
            return row  # بازگرداندن شیء ردیف برای ویرایش
    return None

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "🎓 به ربات تلگرامی جشن فارغ‌التحصیلی بهمن ۹۷ خوش آمدید\n"
        "لطفاً اسم خود را وارد نمایید:"
    )
    return NAME

async def get_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["name"] = update.message.text.strip()
    await update.message.reply_text("لطفاً کد دانشجویی خود را وارد نمایید:")
    return STUDENT_ID

async def get_student_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    student_id = update.message.text.strip()
    context.user_data["student_id"] = student_id

    row = find_student(student_id)
    if row:
        old_name = row[0].value
        old_guest = row[2].value
        context.user_data["excel_row_idx"] = row[0].row  # نگه داشتن شمارهٔ ردیف
        keyboard = [["بله ✅", "خیر ❌"]]
        await update.message.reply_text(
            f"⚠️ این کد دانشجویی قبلاً ثبت شده است:\n\n"
            f"نام: {old_name}\n"
            f"تعداد همراهان: {old_guest}\n\n"
            "آیا مایل به ویرایش اطلاعات هستید؟",
            reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True),
        )
        return EDIT_CONFIRM

    await update.message.reply_text(
        "لطفاً تعداد افرادی که می‌خواهید همراه خود بیاورید را وارد کنید یا بنویسید «نامشخص»:"
    )
    return GUEST_COUNT

async def edit_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    if "بله" in text:
        await update.message.reply_text("لطفاً نام جدید خود را وارد کنید:")
        return EDIT_NAME
    else:
        await update.message.reply_text("✅ اطلاعات قبلی بدون تغییر باقی ماند.")
        return ConversationHandler.END

async def edit_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["new_name"] = update.message.text.strip()
    await update.message.reply_text(
        "لطفاً تعداد جدید همراهان را وارد کنید یا بنویسید «نامشخص»:"
    )
    return EDIT_GUEST

async def edit_guest(update: Update, context: ContextTypes.DEFAULT_TYPE):
    new_guest = update.message.text.strip()
    row_idx = context.user_data.get("excel_row_idx")
    if row_idx is None:
        await update.message.reply_text("خطا: ردیف پیدا نشد.")
        return ConversationHandler.END

    wb = load_workbook(FILE_NAME)
    ws = wb.active
    ws.cell(row=row_idx, column=1, value=context.user_data["new_name"])
    ws.cell(row=row_idx, column=3, value=new_guest)
    wb.save(FILE_NAME)

    await update.message.reply_text("✅ اطلاعات شما با موفقیت ویرایش شد.")
    return ConversationHandler.END

async def get_guest_count(update: Update, context: ContextTypes.DEFAULT_TYPE):
    guest_count = update.message.text.strip()
    name = context.user_data.get("name")
    student_id = context.user_data.get("student_id")

    # جلوگیری از ثبت دوباره (حساس به همون رشته)
    if find_student(student_id):
        await update.message.reply_text("⚠️ خطا: کد دانشجویی شما در همین لحظه ثبت شده است.")
        return ConversationHandler.END

    wb = load_workbook(FILE_NAME)
    ws = wb.active
    ws.append([name, student_id, guest_count])
    wb.save(FILE_NAME)

    await update.message.reply_text("✅ اطلاعات شما با موفقیت ثبت شد.")
    return ConversationHandler.END

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("❌ عملیات لغو شد.")
    return ConversationHandler.END

def main():
    init_excel()
    TOKEN = os.getenv("BOT_TOKEN")
    if not TOKEN:
        raise RuntimeError("توکن پیدا نشد. لطفاً BOT_TOKEN را در فایل .env یا متغیر محیطی قرار دهید.")
    app = ApplicationBuilder().token(TOKEN).build()

    conv = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_name)],
            STUDENT_ID: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_student_id)],
            GUEST_COUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_guest_count)],
            EDIT_CONFIRM: [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_confirm)],
            EDIT_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_name)],
            EDIT_GUEST: [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_guest)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )

    app.add_handler(conv)
    print("Bot is running...")
    app.run_polling()

if __name__ == "__main__":
    main()
